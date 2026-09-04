"""
Smart Meter Field Tracker
=========================
Backend : streamlit-gsheets-connection  (Google Sheets)
Theme   : Clean White & Light Greys (Field-Optimized)
Security: PIN Protected (30-min inactivity auto-lock)

Google Sheet worksheets required (create these tabs in your Sheet, header row only —
the app creates and appends data automatically):
  Installations       - date, tech_name, location, qty_1ph, qty_3ph
  Inventory            - date, type, qty, mrn, make
  Technicians           - name, phone, aadhar, is_active, login_id
  Locations             - location_name
  UploadedInstallLog    - key, date, time, installer_id, tech_name, location, meter_type
  AnalyticsRaw          - key, date, time, installer_id, hour
"""

import streamlit as st
from streamlit_gsheets import GSheetsConnection
import pandas as pd
from datetime import date, datetime, time as dtime
import urllib.parse
import math
import time
import io
import openpyxl

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Field Meter Tracker",
    page_icon="⚡",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ── Constants ─────────────────────────────────────────────────────────────────
PIN_CODE = "1323"
SESSION_TIMEOUT_SECONDS = 30 * 60  # 30 minutes inactivity
READ_TTL = 30  # seconds — cuts down on redundant Sheets reads
HALF_DAY_CUTOFF = "13:30:00"  # H1 = first install .. 13:30, H2 = 13:30 .. last install

# ── Conditional formatting thresholds ────────────────────────────────────────
# Mirrors the colour rules used in the LoginID_Summary sheet of the MDM export.
# Tune these if your team size / daily targets differ.
CF_GREEN_BG, CF_GREEN_FONT = "#C6EFCE", "#006100"
CF_YELLOW_BG, CF_YELLOW_FONT = "#FFEB9C", "#9C5700"
CF_RED_BG, CF_RED_FONT = "#FFC7CE", "#9C0006"

# Per installer × hour cell (e.g. B3:L13 in the source sheet): <2 red, =2 yellow, >2 green
HOURLY_CELL_THRESHOLD = 2
# Per-installer daily Total column (M3:M13 / Q16:Q26): <10 red, 10-15 neutral, 15-20 yellow, >20 green
INSTALLER_TOTAL_RED_MAX, INSTALLER_TOTAL_YELLOW_MIN, INSTALLER_TOTAL_YELLOW_MAX = 10, 15, 20
# Bottom TOTAL row, per-hour aggregate (B14:L14 / U17:U27): same 4-tier scheme
HOURLY_TOTAL_RED_MAX, HOURLY_TOTAL_YELLOW_MIN, HOURLY_TOTAL_YELLOW_MAX = 10, 15, 20
# Grand total for the day (M14): <150 red, 150-200 yellow, >200 green
GRAND_TOTAL_RED_MAX, GRAND_TOTAL_YELLOW_MAX = 150, 200
# Avg install time in minutes (V32:V42): <20 green (fast), 20-30 yellow, >30 red (slow)
AVG_TIME_GREEN_MAX, AVG_TIME_YELLOW_MAX = 20, 30


def tier_style(v, red_max, yellow_min, yellow_max):
    """4-tier: <red_max red · red_max-yellow_min neutral · yellow_min-yellow_max yellow · >yellow_max green."""
    try:
        v = float(v)
    except Exception:
        return ""
    if v < red_max:
        return f"background-color:{CF_RED_BG};color:{CF_RED_FONT}"
    if v < yellow_min:
        return ""
    if v <= yellow_max:
        return f"background-color:{CF_YELLOW_BG};color:{CF_YELLOW_FONT}"
    return f"background-color:{CF_GREEN_BG};color:{CF_GREEN_FONT}"


def cell_style_3tier(v, mid):
    """3-tier for a single count cell: <mid red · =mid yellow · >mid green."""
    try:
        v = float(v)
    except Exception:
        return ""
    if v < mid:
        return f"background-color:{CF_RED_BG};color:{CF_RED_FONT}"
    if v == mid:
        return f"background-color:{CF_YELLOW_BG};color:{CF_YELLOW_FONT}"
    return f"background-color:{CF_GREEN_BG};color:{CF_GREEN_FONT}"


def avg_time_style(v):
    """Lower avg install time is better: <20 green · 20-30 yellow · >30 red."""
    try:
        v = float(v)
    except Exception:
        return ""
    if v <= 0:
        return ""
    if v < AVG_TIME_GREEN_MAX:
        return f"background-color:{CF_GREEN_BG};color:{CF_GREEN_FONT}"
    if v <= AVG_TIME_YELLOW_MAX:
        return f"background-color:{CF_YELLOW_BG};color:{CF_YELLOW_FONT}"
    return f"background-color:{CF_RED_BG};color:{CF_RED_FONT}"


def style_hourly_table(df: pd.DataFrame, hour_cols):
    """Applies the LoginID_Summary-style colouring: per-cell 3-tier for each
    installer's hour buckets, 4-tier for the Total column, and a matching
    4-tier scheme for the bottom aggregate TOTAL row."""
    def styler(data):
        css = pd.DataFrame("", index=data.index, columns=data.columns)
        for i in data.index:
            is_total_row = str(data.loc[i, "Installer"]).strip().upper() == "TOTAL"
            for h in hour_cols:
                if is_total_row:
                    css.loc[i, h] = tier_style(data.loc[i, h], HOURLY_TOTAL_RED_MAX, HOURLY_TOTAL_YELLOW_MIN, HOURLY_TOTAL_YELLOW_MAX)
                else:
                    css.loc[i, h] = cell_style_3tier(data.loc[i, h], HOURLY_CELL_THRESHOLD)
            if "Total" in data.columns:
                css.loc[i, "Total"] = tier_style(data.loc[i, "Total"], INSTALLER_TOTAL_RED_MAX, INSTALLER_TOTAL_YELLOW_MIN, INSTALLER_TOTAL_YELLOW_MAX)
        return css
    return df.style.apply(styler, axis=None)


def render_colored_metric(label: str, value: int, red_max: int, yellow_max: int):
    """A st.metric look-alike whose background/text colour reflects thresholds
    (mirrors the M14 grand-total cell colouring in the source sheet)."""
    if value < red_max:
        bg, fg = CF_RED_BG, CF_RED_FONT
    elif value <= yellow_max:
        bg, fg = CF_YELLOW_BG, CF_YELLOW_FONT
    else:
        bg, fg = CF_GREEN_BG, CF_GREEN_FONT
    st.markdown(f"""
    <div style="background:{bg};color:{fg};border-radius:14px;padding:16px 14px;
        border:1px solid rgba(0,0,0,0.06);text-align:left;">
        <div style="font-size:.72rem;font-weight:600;text-transform:uppercase;letter-spacing:.4px;opacity:.85;">{label}</div>
        <div style="font-size:1.7rem;font-weight:800;letter-spacing:-.3px;">{value}</div>
    </div>
    """, unsafe_allow_html=True)


# ── CSS – Fintech-Inspired Theme (single accent, segmented tabs) ────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

:root {
    --accent: #0E9F6E;
    --accent-dark: #0B7A56;
    --accent-soft: #E6F7F0;
    --ink: #10151F;
    --ink-soft: #64748B;
    --bg: #F6F7F9;
    --card-border: #E7E9EE;
}

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: var(--bg); color: var(--ink); }
#MainMenu, footer, header { visibility:hidden; }

.top-banner {
    background: #ffffff;
    border: 1px solid var(--card-border);
    border-radius: 16px;
    padding: 14px 18px;
    display:flex; align-items:center; gap:12px;
    box-shadow: 0 1px 2px rgba(16,21,31,0.04);
}
.top-banner .icon-badge {
    width:40px; height:40px; border-radius:12px; background:var(--accent-soft);
    display:flex; align-items:center; justify-content:center; font-size:1.3rem; flex-shrink:0;
}
.top-banner .t { font-size:1.15rem; font-weight:800; color:var(--ink); letter-spacing:-.2px; margin:0; }
.top-banner .s { font-size:.78rem; color:var(--ink-soft); margin:0; font-weight:500; }

/* Segmented-control style tabs, closer to Groww/Kite bottom-nav feel */
.stTabs [data-baseweb="tab-list"] {
    background:#EEF0F3; border-radius:12px; padding:4px; gap:2px;
    overflow-x:auto; white-space:nowrap;
}
.stTabs [data-baseweb="tab"] {
    border-radius:9px !important; padding:9px 14px !important;
    font-size:.86rem !important; font-weight:600 !important;
    color:var(--ink-soft) !important;
    background:transparent !important; border:none !important;
}
.stTabs [aria-selected="true"] {
    background:var(--ink) !important;
    color:#ffffff !important;
    box-shadow: 0 1px 3px rgba(16,21,31,0.15);
}

[data-testid="stMetric"] {
    background: #ffffff;
    border: 1px solid var(--card-border); border-radius:14px;
    padding: 16px 14px !important;
    box-shadow: 0 1px 2px rgba(16,21,31,0.03);
}
[data-testid="stMetricLabel"] {
    color:var(--ink-soft) !important; font-size:.72rem !important; font-weight:600 !important;
    text-transform:uppercase; letter-spacing:.4px;
}
[data-testid="stMetricValue"] {
    font-size:1.7rem !important; font-weight:800 !important; color:var(--ink) !important; letter-spacing:-.3px;
}

.sec-hdr {
    font-size:1.02rem; font-weight:700; color:var(--ink);
    display:flex; align-items:center; gap:8px;
    margin: 1.6rem 0 .9rem;
}
.sec-hdr::before { content:""; width:5px; height:16px; background:var(--accent); border-radius:3px; display:inline-block; }
.sub-hdr {
    font-size:.85rem; font-weight:700; color:var(--ink-soft);
    text-transform:uppercase; letter-spacing:.4px;
    margin: 1.1rem 0 .5rem;
}

.stButton>button {
    background:#ffffff !important; color:var(--ink) !important;
    border:1px solid var(--card-border) !important; border-radius:10px !important;
    font-weight:600 !important; font-size:.92rem !important;
    padding:10px 18px !important; width:100% !important;
    transition:all .15s;
    box-shadow: 0 1px 2px rgba(16,21,31,0.02);
}
.stButton>button:hover { border-color:var(--accent) !important; color:var(--accent-dark) !important; }

button[data-testid="baseButton-primary"], .stButton>button[type="primary"] {
    background:var(--accent) !important; color:#ffffff !important; border-color:var(--accent) !important;
}
button[data-testid="baseButton-primary"]:hover, .stButton>button[type="primary"]:hover {
    background:var(--accent-dark) !important; border-color:var(--accent-dark) !important; color:#fff !important;
}

.stSelectbox>div>div, .stNumberInput>div>div>input,
.stTextInput>div>div>input, .stDateInput>div>div>input, .stMultiSelect>div>div {
    background:#ffffff !important; border:1px solid var(--card-border) !important;
    border-radius:10px !important; color:var(--ink) !important; font-size:.9rem !important;
}

.stForm { background:#ffffff !important; border:1px solid var(--card-border) !important;
    border-radius:14px !important; padding:18px !important; }

.stDataFrame { border-radius:12px; border: 1px solid var(--card-border); overflow:hidden; }

.warn-box {
    background:#FFF8E8; border:1px solid #F5D98B; border-radius:11px;
    padding:11px 15px; color:#8A6208; font-size:.85rem; margin-bottom:.8rem; font-weight:500;
}
.info-box {
    background:#F1F5F9; border:1px solid var(--card-border); border-radius:11px;
    padding:11px 15px; color:var(--ink-soft); font-size:.85rem; margin-bottom:.8rem; font-weight:500;
}
.danger-box {
    background:#FEF2F2; border:1px solid #FCA5A5; border-radius:11px;
    padding:11px 15px; color:#991B1B; font-size:.85rem; margin-bottom:.8rem; font-weight:500;
}

.wa-btn {
    display:block; text-align:center; background:#25D366; color:#fff !important;
    padding:13px; border-radius:11px; text-decoration:none; font-weight:700;
    font-size:1rem; letter-spacing:.2px;
    margin-top:1rem; transition: background 0.2s;
    box-shadow: 0 2px 6px rgba(37,211,102,0.25);
}
.wa-btn:hover { background:#1DA851; }
</style>
""", unsafe_allow_html=True)


# ── Top banner & Refresh Button ───────────────────────────────────────────────
head_col1, head_col2 = st.columns([3.5, 1.2])
with head_col1:
    st.markdown("""
    <div class="top-banner">
      <div class="icon-badge">⚡</div>
      <div>
        <p class="t">Meter Tracker</p>
        <p class="s">Vijayawada Field Ops</p>
      </div>
    </div>
    """, unsafe_allow_html=True)
with head_col2:
    st.write("")
    if st.button("🔄 Refresh"):
        st.cache_data.clear()
        st.rerun()

st.write("")

# ── Authentication / PIN Protection (30-min inactivity auto-lock) ────────────
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
if "last_activity" not in st.session_state:
    st.session_state["last_activity"] = time.time()

# If session has gone stale, force re-login before rendering anything else.
if st.session_state["authenticated"]:
    idle_for = time.time() - st.session_state["last_activity"]
    if idle_for > SESSION_TIMEOUT_SECONDS:
        st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.markdown('<div class="sec-hdr">🔒 Supervisor Login</div>', unsafe_allow_html=True)
    with st.form("login_form"):
        st.info("Please enter the daily operations PIN to access the system. You'll stay logged in for 30 minutes of inactivity.")
        pin_entry = st.text_input("Enter PIN", type="password")
        login_btn = st.form_submit_button("Unlock Tracker", type="primary")
        if login_btn:
            if pin_entry == PIN_CODE:
                st.session_state["authenticated"] = True
                st.session_state["last_activity"] = time.time()
                st.success("Access Granted!")
                st.rerun()
            else:
                st.error("❌ Incorrect PIN. Access Denied.")
    st.stop()

# Any render past this point means an active, valid session — refresh the clock.
st.session_state["last_activity"] = time.time()

# ── Cloud Crash Guard: Google Sheets Connection ──────────────────────────────
try:
    conn = st.connection("gsheets", type=GSheetsConnection)
except Exception as e:
    st.error("🛑 Database Connection Failed!")
    st.write(f"Error Details: `{e}`")
    st.info("💡 **Fix:** Ensure your `st.secrets` are properly configured and `st-gsheets-connection` is in requirements.txt")
    st.stop()

# ── Helpers ───────────────────────────────────────────────────────────────────
def get_data(worksheet: str, retries: int = 3) -> pd.DataFrame:
    for attempt in range(retries):
        try:
            df = conn.read(worksheet=worksheet, ttl=READ_TTL)
            return df.astype(str).fillna("") if not df.empty else pd.DataFrame()
        except Exception:
            if attempt < retries - 1:
                time.sleep(1)
            else:
                st.toast(f"📡 Connection drop loading {worksheet}...", icon="⚠️")
                return pd.DataFrame()


def safe_update(worksheet: str, data: pd.DataFrame, retries: int = 3) -> bool:
    """Write to Sheets with retries so a dropped connection doesn't lose the entry.
    On repeated failure, the data the user entered is NOT cleared — they can just retry."""
    for attempt in range(retries):
        try:
            with st.spinner(f"💾 Saving to {worksheet}..."):
                conn.update(worksheet=worksheet, data=data.astype(str))
            st.cache_data.clear()
            return True
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1.5)
            else:
                st.error(f"⚠️ Save failed after several attempts ({e}). Your entries are still in the form — please tap Save again once you have signal.")
                return False
    return False


def safe_int(val, default: int = 0) -> int:
    try:
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return default
        return int(float(val))
    except Exception:
        return default


def safe_numeric_col(df: pd.DataFrame, col: str) -> pd.Series:
    return pd.to_numeric(df[col], errors="coerce").fillna(0)


def has_col(df: pd.DataFrame, *cols) -> bool:
    return all(c in df.columns for c in cols)


# ── Excel parsing helpers (used by Installs bulk upload + Analytics upload) ─
def find_header_row(ws, required_headers, max_scan_rows: int = 20):
    """Scan the first N rows for a row containing all required header labels
    (case-insensitive, trimmed). Returns (row_index, {header_label: col_index})
    or (None, None) if not found. Works regardless of whether headers sit on
    row 1 (clean export) or a later row (raw MDM export with a title row)."""
    required_norm = [h.strip().lower() for h in required_headers]
    max_row = min(max_scan_rows, ws.max_row)
    for r in range(1, max_row + 1):
        row_vals = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                row_vals[str(v).strip().lower()] = c
        if all(h in row_vals for h in required_norm):
            col_map = {orig: row_vals[norm] for orig, norm in zip(required_headers, required_norm)}
            return r, col_map
    return None, None


def normalize_date_val(val):
    """Return an ISO date string (YYYY-MM-DD) or None."""
    if val is None:
        return None
    try:
        if isinstance(val, datetime):
            return val.date().isoformat()
        if isinstance(val, date):
            return val.isoformat()
        s = str(val).strip()
        if not s:
            return None
        parsed = pd.to_datetime(s, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date().isoformat()
    except Exception:
        return None


def normalize_time_val(val):
    """Return a zero-padded HH:MM:SS string or None."""
    if val is None:
        return None
    try:
        if isinstance(val, datetime):
            return val.strftime("%H:%M:%S")
        if isinstance(val, dtime):
            return val.strftime("%H:%M:%S")
        s = str(val).strip()
        if not s:
            return None
        parsed = pd.to_datetime(s, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.strftime("%H:%M:%S")
    except Exception:
        return None


def load_first_data_sheet(uploaded_file):
    """Return the primary data worksheet from an uploaded workbook, skipping
    any pre-computed pivot/summary sheets (e.g. 'LoginID_Summary')."""
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    for name in wb.sheetnames:
        if "summary" not in name.strip().lower():
            return wb[name]
    return wb[wb.sheetnames[0]]


def time_to_minutes(hhmmss: str) -> float:
    h, m, s = hhmmss.split(":")
    return int(h) * 60 + int(m) + int(s) / 60.0


# ── Shared data fetched once per run (avoids repeat reads across tabs) ──────
df_installations_master = get_data("Installations")
df_inventory_master = get_data("Inventory")
df_technicians_master = get_data("Technicians")
df_locations_master = get_data("Locations")

active_techs = []
if not df_technicians_master.empty and has_col(df_technicians_master, "is_active", "name"):
    for _, r in df_technicians_master.iterrows():
        if str(r["is_active"]).strip().lower() in ["1", "1.0", "true", "yes"]:
            n = str(r["name"]).strip()
            if n:
                active_techs.append(n)

active_locs = []
if not df_locations_master.empty and "location_name" in df_locations_master.columns:
    for _, r in df_locations_master.iterrows():
        l = str(r["location_name"]).strip()
        if l:
            active_locs.append(l)

# Installer LoginID -> Technician display name, from the optional "login_id"
# column on the Technicians sheet. Unmapped logins fall back to the raw ID.
tech_login_lookup = {}
if not df_technicians_master.empty and has_col(df_technicians_master, "login_id", "name"):
    for _, r in df_technicians_master.iterrows():
        lid = str(r.get("login_id", "")).strip().lower()
        nm = str(r.get("name", "")).strip()
        if lid and nm:
            tech_login_lookup[lid] = nm

# ── Tabs Configuration ────────────────────────────────────────────────────────
tab_dash, tab_analytics, tab_inst, tab_inv, tab_admin = st.tabs([
    "📊 Dashboard", "📈 Analytics", "🛠️ Installs", "📦 Store", "⚙️ Admin"
])

# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
with tab_dash:
    df_inst = df_installations_master
    df_inv = df_inventory_master

    st.markdown('<div class="sec-hdr">📦 Live Inventory Stock</div>', unsafe_allow_html=True)

    if not df_inv.empty and has_col(df_inv, "type", "qty"):
        total_in_1ph = safe_numeric_col(df_inv[df_inv["type"] == "1 PH"], "qty").sum()
        total_in_3ph = safe_numeric_col(df_inv[df_inv["type"] == "3 PH"], "qty").sum()
    else:
        total_in_1ph = total_in_3ph = 0

    if not df_inst.empty and has_col(df_inst, "qty_1ph", "qty_3ph"):
        total_out_1ph = safe_numeric_col(df_inst, "qty_1ph").sum()
        total_out_3ph = safe_numeric_col(df_inst, "qty_3ph").sum()
    else:
        total_out_1ph = total_out_3ph = 0

    pending_1ph = int(total_in_1ph - total_out_1ph)
    pending_3ph = int(total_in_3ph - total_out_3ph)

    sc1, sc2, sc3, sc4 = st.columns(4)
    sc1.metric("Received 1PH", int(total_in_1ph))
    sc2.metric("Received 3PH", int(total_in_3ph))
    sc3.metric("Pending 1PH", pending_1ph, delta="⚠️ Deficit!" if pending_1ph < 0 else None, delta_color="inverse")
    sc4.metric("Pending 3PH", pending_3ph, delta="⚠️ Deficit!" if pending_3ph < 0 else None, delta_color="inverse")

    # ── Monthly Installs Overview ────────────────────────────────────────────
    st.divider()
    st.markdown('<div class="sec-hdr">📅 Monthly Installs Overview</div>', unsafe_allow_html=True)

    if df_inst.empty or not has_col(df_inst, "date", "qty_1ph", "qty_3ph", "location"):
        st.info("No installation data yet.")
    else:
        df_month = df_inst.copy()
        df_month["_date"] = pd.to_datetime(df_month["date"], errors="coerce")
        df_month["qty_1ph"] = safe_numeric_col(df_month, "qty_1ph")
        df_month["qty_3ph"] = safe_numeric_col(df_month, "qty_3ph")

        today = date.today()
        this_month = df_month[(df_month["_date"].dt.month == today.month) & (df_month["_date"].dt.year == today.year)]

        tm1, tm2, tm3 = st.columns(3)
        tm1.metric("This Month — 1PH", int(this_month["qty_1ph"].sum()))
        tm2.metric("This Month — 3PH", int(this_month["qty_3ph"].sum()))
        tm3.metric("This Month — Total", int(this_month["qty_1ph"].sum() + this_month["qty_3ph"].sum()))

        st.markdown('<div class="sub-hdr">📍 This Month, By Location</div>', unsafe_allow_html=True)
        if this_month.empty:
            st.info("No installs recorded this month yet.")
        else:
            loc_month = this_month.groupby("location")[["qty_1ph", "qty_3ph"]].sum().reset_index()
            loc_month["Total"] = loc_month["qty_1ph"] + loc_month["qty_3ph"]
            loc_month.columns = ["Location", "1PH", "3PH", "Total"]
            loc_month = loc_month.sort_values("Total", ascending=False)
            st.dataframe(loc_month, use_container_width=True, hide_index=True)

    st.divider()
    st.markdown('<div class="sec-hdr">🔌 Installation Summary (Filterable)</div>', unsafe_allow_html=True)

    if df_inst.empty or not has_col(df_inst, "date", "tech_name", "location", "qty_1ph", "qty_3ph"):
        st.info("No installation data yet. Add entries in the Installs tab.")
    else:
        f1, f2 = st.columns(2)
        with f1:
            date_range = st.date_input("Date Range", [date.today(), date.today()])
        with f2:
            meter_filter = st.multiselect("Meter Type", ["1 PH", "3 PH"], default=["1 PH", "3 PH"])

        loc_list = sorted([l for l in df_inst["location"].unique() if l.strip()])
        tech_list = sorted([t for t in df_inst["tech_name"].unique() if t.strip()])

        f3, f4 = st.columns(2)
        with f3:
            loc_filter = st.multiselect("Locations", loc_list, default=loc_list)
        with f4:
            tech_filter = st.multiselect("Technicians", tech_list, default=tech_list)

        filtered = df_inst.copy()

        if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
            d_start, d_end = date_range[0], date_range[1]
        elif isinstance(date_range, (list, tuple)) and len(date_range) == 1:
            d_start = d_end = date_range[0]
        else:
            d_start = d_end = date_range

        filtered["_date"] = pd.to_datetime(filtered["date"], errors="coerce").dt.date
        filtered = filtered[(filtered["_date"] >= d_start) & (filtered["_date"] <= d_end)]
        if loc_filter:
            filtered = filtered[filtered["location"].isin(loc_filter)]
        if tech_filter:
            filtered = filtered[filtered["tech_name"].isin(tech_filter)]

        filtered["qty_1ph"] = safe_numeric_col(filtered, "qty_1ph")
        filtered["qty_3ph"] = safe_numeric_col(filtered, "qty_3ph")

        show_1ph, show_3ph = "1 PH" in meter_filter, "3 PH" in meter_filter
        sum_1ph = int(filtered["qty_1ph"].sum()) if show_1ph else 0
        sum_3ph = int(filtered["qty_3ph"].sum()) if show_3ph else 0

        m1, m2, m3 = st.columns(3)
        m1.metric("Filtered 1PH", sum_1ph)
        m2.metric("Filtered 3PH", sum_3ph)
        m3.metric("Grand Total", sum_1ph + sum_3ph)

        if not filtered.empty:
            st.markdown('<div class="sec-hdr">👷 Technician Breakdown</div>', unsafe_allow_html=True)
            group_df = filtered.groupby(["tech_name", "location"])[["qty_1ph", "qty_3ph"]].sum().reset_index()
            group_df["Total"] = group_df["qty_1ph"] + group_df["qty_3ph"]
            group_df.columns = ["Technician", "Location", "1PH", "3PH", "Total"]
            st.dataframe(
                group_df.style.apply(
                    lambda data: pd.DataFrame(
                        {c: (data["Total"].apply(lambda v: tier_style(v, INSTALLER_TOTAL_RED_MAX, INSTALLER_TOTAL_YELLOW_MIN, INSTALLER_TOTAL_YELLOW_MAX)) if c == "Total" else "") for c in data.columns},
                        index=data.index,
                    ),
                    axis=None,
                ),
                use_container_width=True, hide_index=True,
            )
            st.caption("🟩 Green = strong Total · 🟨 Yellow = mid-range · 🟥 Red = below target.")

            st.markdown('<div class="sec-hdr">📤 Export & Share</div>', unsafe_allow_html=True)
            export_df = group_df.copy()
            export_df.loc[len(export_df)] = ["---", "---", "---", "---", "---"]
            export_df.loc[len(export_df)] = ["GRAND TOTAL", "", sum_1ph, sum_3ph, sum_1ph + sum_3ph]
            export_df.loc[len(export_df)] = ["PENDING STOCK", "", pending_1ph, pending_3ph, ""]

            csv_data = export_df.to_csv(index=False).encode("utf-8")
            st.download_button("📥 Download CSV Report", data=csv_data, file_name="Installation_Summary.csv", mime="text/csv", use_container_width=True)

            date_str = f"{d_start} to {d_end}" if d_start != d_end else str(d_start)
            wa_loc_df = filtered.groupby("location")[["qty_1ph", "qty_3ph"]].sum().reset_index()

            wa_lines = ["DPR- Touchlight Infra", f"Date: {date_str}\n"]
            for _, row in wa_loc_df.iterrows():
                wa_lines.append(f"{row['location']}:")
                wa_lines.append(f"1PH: {int(row['qty_1ph']) if show_1ph else 0}, 3PH: {int(row['qty_3ph']) if show_3ph else 0}\n")

            wa_lines.append(f"Total 1PH: {sum_1ph} | Total 3PH: {sum_3ph} | Grand Total: {sum_1ph + sum_3ph}")
            wa_lines.append(f"Pending Stock: 1PH: {pending_1ph} | 3PH: {pending_3ph}")

            wa_text = "\n".join(wa_lines)
            wa_url = f"https://wa.me/?text={urllib.parse.quote(wa_text)}"
            st.markdown(f'<a href="{wa_url}" target="_blank" class="wa-btn">💬 Send to WhatsApp</a>', unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
#  ANALYTICS  (fully independent of Installations/Inventory/Technicians —
#  purely for live installer-performance tracking on the phone while traveling)
# ═══════════════════════════════════════════════════════════════════════════════
with tab_analytics:
    st.markdown("""
    <div class="info-box">
    📈 This tab is independent of the Installs/Inventory data elsewhere in the app.
    Upload the raw MDM export (any layout — the app finds the header row automatically)
    to see live installer-wise hourly counts, half-day split, and average install time,
    even when you don't have laptop access. Uploading the same file again only adds
    genuinely new rows — nothing is double counted. Reset at the end of the day to start fresh.
    </div>
    """, unsafe_allow_html=True)

    ANALYTICS_REQUIRED_HEADERS = ["Installation Date", "Installation Time", "Installer LoginID"]

    st.markdown('<div class="sec-hdr">⬆️ Upload Progress File</div>', unsafe_allow_html=True)
    analytics_file = st.file_uploader(
        "Upload the MDM export (.xlsx) — only Installer LoginIDs starting with TL_ are counted",
        type=["xlsx"], key="analytics_uploader"
    )

    if analytics_file is not None:
        if st.button("📊 Process & Add To Analytics", type="primary", use_container_width=True):
            try:
                ws = load_first_data_sheet(analytics_file)
            except Exception as e:
                st.error(f"❌ Could not open the file: {e}")
                ws = None

            if ws is not None:
                header_row, col_map = find_header_row(ws, ANALYTICS_REQUIRED_HEADERS)
                if header_row is None:
                    st.error("❌ Could not find 'Installation Date', 'Installation Time' and 'Installer LoginID' columns in this file.")
                else:
                    parsed_records = []
                    skipped_non_tl = 0
                    for r in range(header_row + 1, ws.max_row + 1):
                        raw_installer = ws.cell(row=r, column=col_map["Installer LoginID"]).value
                        if raw_installer is None or str(raw_installer).strip() == "":
                            continue
                        installer_id = str(raw_installer).strip()
                        if not installer_id.upper().startswith("TL_"):
                            skipped_non_tl += 1
                            continue
                        d = normalize_date_val(ws.cell(row=r, column=col_map["Installation Date"]).value)
                        t = normalize_time_val(ws.cell(row=r, column=col_map["Installation Time"]).value)
                        if d is None or t is None:
                            continue
                        parsed_records.append({
                            "date": d, "time": t, "installer_id": installer_id,
                            "hour": t.split(":")[0],
                        })

                    if not parsed_records:
                        st.warning("⚠️ No valid TL_ installer rows with a date and time were found in this file.")
                    else:
                        df_araw_existing = get_data("AnalyticsRaw")
                        existing_keys = set()
                        if not df_araw_existing.empty and "key" in df_araw_existing.columns:
                            existing_keys = set(df_araw_existing["key"].values)

                        new_rows = []
                        dup_count = 0
                        for rec in parsed_records:
                            key = f"{rec['date']}||{rec['time']}||{rec['installer_id']}"
                            if key in existing_keys:
                                dup_count += 1
                                continue
                            existing_keys.add(key)
                            new_rows.append({
                                "key": key, "date": rec["date"], "time": rec["time"],
                                "installer_id": rec["installer_id"], "hour": rec["hour"],
                            })

                        if not new_rows:
                            st.error("❌ All records in this file are already in Analytics (duplicate date/time/installer). Nothing new to add.")
                        else:
                            merged = pd.concat([df_araw_existing, pd.DataFrame(new_rows)], ignore_index=True) if not df_araw_existing.empty else pd.DataFrame(new_rows)
                            if safe_update("AnalyticsRaw", merged):
                                msg = f"✅ Added {len(new_rows)} new record(s) to Analytics."
                                if dup_count:
                                    msg += f" Skipped {dup_count} already-recorded duplicate(s)."
                                if skipped_non_tl:
                                    msg += f" Ignored {skipped_non_tl} non-TL_ installer row(s)."
                                st.success(msg)
                                st.rerun()

    # ── Build analytics tables from stored raw data ─────────────────────────
    st.divider()
    df_araw = get_data("AnalyticsRaw")

    if df_araw.empty or not has_col(df_araw, "date", "time", "installer_id", "hour"):
        st.info("No analytics data yet — upload a progress file above to get started.")
    else:
        avail_dates = sorted(df_araw["date"].unique(), reverse=True)
        sel_date = st.selectbox("Viewing date", avail_dates, index=0)
        day_df = df_araw[df_araw["date"] == sel_date].copy()
        day_df["hour_int"] = pd.to_numeric(day_df["hour"], errors="coerce")
        installers = sorted(day_df["installer_id"].unique())

        st.markdown('<div class="sec-hdr">📌 Today At A Glance</div>', unsafe_allow_html=True)
        g1, g2, g3 = st.columns(3)
        with g1:
            render_colored_metric("Total Installs", len(day_df), GRAND_TOTAL_RED_MAX, GRAND_TOTAL_YELLOW_MAX)
        g2.metric("Active Installers", len(installers))
        g3.metric("Avg / Installer", round(len(day_df) / len(installers), 1) if installers else 0)

        # -- Hourly table --------------------------------------------------
        st.markdown('<div class="sec-hdr">⏱️ Installer-Wise Hourly Count</div>', unsafe_allow_html=True)
        if day_df["hour_int"].notna().any():
            hr_min = int(day_df["hour_int"].min())
            hr_max = int(day_df["hour_int"].max())
        else:
            hr_min, hr_max = 8, 18

        hour_cols = list(range(hr_min, hr_max + 1))
        hourly_rows = []
        for inst in installers:
            sub = day_df[day_df["installer_id"] == inst]
            row = {"Installer": inst}
            for h in hour_cols:
                row[f"{h}-{h+1}"] = int((sub["hour_int"] == h).sum())
            row["Total"] = len(sub)
            hourly_rows.append(row)
        hourly_df = pd.DataFrame(hourly_rows).sort_values("Total", ascending=False)
        total_row = {"Installer": "TOTAL"}
        for h in hour_cols:
            total_row[f"{h}-{h+1}"] = int(hourly_df[f"{h}-{h+1}"].sum())
        total_row["Total"] = int(hourly_df["Total"].sum())
        hourly_df = pd.concat([hourly_df, pd.DataFrame([total_row])], ignore_index=True)
        hour_col_labels = [f"{h}-{h+1}" for h in hour_cols]
        st.dataframe(style_hourly_table(hourly_df, hour_col_labels), use_container_width=True, hide_index=True)
        st.caption("🟩 Green = strong count · 🟨 Yellow = mid-range · 🟥 Red = below target — thresholds set in the code's Conditional formatting section.")

        # -- Half-day split --------------------------------------------------
        st.markdown('<div class="sec-hdr">🌓 Half-Day Split (H1: start – 13:30 · H2: 13:30 – end)</div>', unsafe_allow_html=True)
        half_rows = []
        for inst in installers:
            sub = day_df[day_df["installer_id"] == inst]
            h1 = int((sub["time"] <= HALF_DAY_CUTOFF).sum())
            h2 = int((sub["time"] > HALF_DAY_CUTOFF).sum())
            half_rows.append({"Installer": inst, "H1 (Morning)": h1, "H2 (Afternoon)": h2, "Total": h1 + h2})
        half_df = pd.DataFrame(half_rows).sort_values("Total", ascending=False)
        st.dataframe(half_df, use_container_width=True, hide_index=True)

        # -- Average install time -------------------------------------------
        st.markdown('<div class="sec-hdr">⏳ Average Install Time / Installer</div>', unsafe_allow_html=True)
        st.caption("Avg (min) = (last install time − first install time in minutes) ÷ total installs for that installer")
        avg_rows = []
        for inst in installers:
            sub = day_df[day_df["installer_id"] == inst].sort_values("time")
            first_t, last_t = sub["time"].iloc[0], sub["time"].iloc[-1]
            n = len(sub)
            span_min = time_to_minutes(last_t) - time_to_minutes(first_t)
            avg_min = round(span_min / n, 1) if n > 0 else 0
            avg_rows.append({
                "Installer": inst, "First Install": first_t, "Last Install": last_t,
                "Total Installs": n, "Avg Time/Install (min)": avg_min,
            })
        avg_df = pd.DataFrame(avg_rows).sort_values("Total Installs", ascending=False)
        st.dataframe(
            avg_df.style.applymap(avg_time_style, subset=["Avg Time/Install (min)"]),
            use_container_width=True, hide_index=True,
        )
        st.caption("🟩 Faster than target · 🟨 Mid-range · 🟥 Slower than target (lower minutes is better).")

        # -- Quick visual ------------------------------------------------------
        st.markdown('<div class="sec-hdr">📊 Total Installs By Installer</div>', unsafe_allow_html=True)
        chart_df = half_df.set_index("Installer")[["Total"]]
        st.bar_chart(chart_df)

        # -- Locked reset --------------------------------------------------
        st.divider()
        with st.expander("🔒 Reset Analytics Data (start a new day)"):
            st.markdown('<div class="danger-box">⚠️ This permanently deletes all Analytics data collected so far. Do this at the end of the day, once you\'re done reviewing.</div>', unsafe_allow_html=True)
            reset_pin = st.text_input("Enter PIN to unlock reset", type="password", key="analytics_reset_pin")
            if reset_pin == PIN_CODE:
                confirm_reset = st.checkbox("I understand this will delete all Analytics data collected so far")
                if st.button("🗑️ Reset Analytics Data", type="primary", disabled=not confirm_reset, use_container_width=True):
                    empty_df = pd.DataFrame(columns=["key", "date", "time", "installer_id", "hour"])
                    if safe_update("AnalyticsRaw", empty_df):
                        st.success("✅ Analytics data cleared. Ready for a new day.")
                        st.rerun()
            elif reset_pin:
                st.error("❌ Incorrect PIN.")

# ═══════════════════════════════════════════════════════════════════════════════
#  INSTALLS
# ═══════════════════════════════════════════════════════════════════════════════
with tab_inst:
    # ── Bulk Upload from MDM Excel export ────────────────────────────────────
    st.markdown('<div class="sec-hdr">📤 Bulk Upload From Excel</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="info-box">
    Upload the daily installation export instead of entering counts manually.
    The app matches each row's <b>Installer LoginID</b>, <b>Date</b>, <b>Time</b> and
    <b>New Meter Type</b>, counts by <b>Section</b> (used as Location), and skips anything
    already saved — so uploading the same file twice won't double-count. If a new file has
    a few extra rows for a date you've already uploaded, only the new ones get added.
    </div>
    """, unsafe_allow_html=True)

    INSTALL_BULK_REQUIRED_HEADERS = ["Installation Date", "Installation Time", "Installer LoginID", "Section", "New Meter Type"]

    bulk_file = st.file_uploader("Upload Installation Excel (.xlsx)", type=["xlsx"], key="bulk_install_uploader")

    if bulk_file is not None:
        if st.button("📥 Process & Save Installs", type="primary", use_container_width=True):
            try:
                ws = load_first_data_sheet(bulk_file)
            except Exception as e:
                st.error(f"❌ Could not open the file: {e}")
                ws = None

            if ws is not None:
                header_row, col_map = find_header_row(ws, INSTALL_BULK_REQUIRED_HEADERS)
                if header_row is None:
                    st.error("❌ Could not find 'Installation Date', 'Installation Time', 'Installer LoginID', 'Section' and 'New Meter Type' columns in this file.")
                else:
                    parsed = []
                    for r in range(header_row + 1, ws.max_row + 1):
                        raw_installer = ws.cell(row=r, column=col_map["Installer LoginID"]).value
                        if raw_installer is None or str(raw_installer).strip() == "":
                            continue
                        d = normalize_date_val(ws.cell(row=r, column=col_map["Installation Date"]).value)
                        t = normalize_time_val(ws.cell(row=r, column=col_map["Installation Time"]).value)
                        if d is None or t is None:
                            continue
                        section = ws.cell(row=r, column=col_map["Section"]).value
                        mtype = ws.cell(row=r, column=col_map["New Meter Type"]).value
                        parsed.append({
                            "date": d, "time": t,
                            "installer_id": str(raw_installer).strip(),
                            "location": str(section).strip() if section else "Unspecified",
                            "meter_type": str(mtype).strip() if mtype else "",
                        })

                    if not parsed:
                        st.warning("⚠️ No valid rows with a date, time and installer were found in this file.")
                    else:
                        df_log_existing = get_data("UploadedInstallLog")
                        existing_keys = set()
                        if not df_log_existing.empty and "key" in df_log_existing.columns:
                            existing_keys = set(df_log_existing["key"].values)

                        new_log_rows = []
                        dates_seen, dates_with_new, unmapped_ids = set(), set(), set()
                        for rec in parsed:
                            dates_seen.add(rec["date"])
                            key = f"{rec['date']}||{rec['time']}||{rec['installer_id']}"
                            if key in existing_keys:
                                continue
                            existing_keys.add(key)
                            tech_name = tech_login_lookup.get(rec["installer_id"].lower())
                            if tech_name is None:
                                tech_name = rec["installer_id"]
                                unmapped_ids.add(rec["installer_id"])
                            new_log_rows.append({
                                "key": key, "date": rec["date"], "time": rec["time"],
                                "installer_id": rec["installer_id"], "tech_name": tech_name,
                                "location": rec["location"], "meter_type": rec["meter_type"],
                            })
                            dates_with_new.add(rec["date"])

                        fully_dup_dates = dates_seen - dates_with_new

                        if not new_log_rows:
                            st.error(f"❌ Installs already exist for: {', '.join(sorted(dates_seen))}. Nothing new to add.")
                        else:
                            # 1) append raw log rows (dedup ledger)
                            updated_log = pd.concat([df_log_existing, pd.DataFrame(new_log_rows)], ignore_index=True) if not df_log_existing.empty else pd.DataFrame(new_log_rows)

                            # 2) aggregate the NEW rows only, by date + tech_name + location
                            new_log_df = pd.DataFrame(new_log_rows)
                            new_log_df["is_1ph"] = new_log_df["meter_type"].str.contains("1", na=False)
                            new_log_df["is_3ph"] = new_log_df["meter_type"].str.contains("3", na=False)
                            agg = new_log_df.groupby(["date", "tech_name", "location"]).agg(
                                d_1ph=("is_1ph", "sum"), d_3ph=("is_3ph", "sum")
                            ).reset_index()

                            # 3) merge deltas into Installations sheet
                            df_inst_existing = get_data("Installations")
                            if df_inst_existing.empty:
                                df_inst_existing = pd.DataFrame(columns=["date", "tech_name", "location", "qty_1ph", "qty_3ph"])
                            for col in ["qty_1ph", "qty_3ph"]:
                                if col in df_inst_existing.columns:
                                    df_inst_existing[col] = pd.to_numeric(df_inst_existing[col], errors="coerce").fillna(0).astype(int)

                            for _, arow in agg.iterrows():
                                mask = (
                                    (df_inst_existing.get("date") == arow["date"]) &
                                    (df_inst_existing.get("tech_name") == arow["tech_name"]) &
                                    (df_inst_existing.get("location") == arow["location"])
                                ) if not df_inst_existing.empty else pd.Series([], dtype=bool)
                                if not df_inst_existing.empty and mask.any():
                                    df_inst_existing.loc[mask, "qty_1ph"] += int(arow["d_1ph"])
                                    df_inst_existing.loc[mask, "qty_3ph"] += int(arow["d_3ph"])
                                else:
                                    df_inst_existing = pd.concat([df_inst_existing, pd.DataFrame([{
                                        "date": arow["date"], "tech_name": arow["tech_name"], "location": arow["location"],
                                        "qty_1ph": int(arow["d_1ph"]), "qty_3ph": int(arow["d_3ph"]),
                                    }])], ignore_index=True)

                            if safe_update("Installations", df_inst_existing) and safe_update("UploadedInstallLog", updated_log):
                                msg = f"✅ Added {len(new_log_rows)} new install(s) across {len(dates_with_new)} date(s)."
                                st.success(msg)
                                if fully_dup_dates:
                                    st.warning(f"⚠️ Already fully recorded, skipped: {', '.join(sorted(fully_dup_dates))}")
                                if unmapped_ids:
                                    st.info(f"ℹ️ No technician mapping found for: {', '.join(sorted(unmapped_ids))} — used their login ID as the name. Add a 'login_id' to that technician in Admin to map it to a display name next time.")
                                st.rerun()

    st.divider()
    st.markdown('<div class="sec-hdr">➕ Daily Entry (Add Multiple At Once)</div>', unsafe_allow_html=True)

    if not active_techs or not active_locs:
        st.warning("⚠️ Please add active Technicians and Locations in the **Admin** tab first.")
    else:
        if "installs_batch" not in st.session_state:
            st.session_state["installs_batch"] = []
        if "qm_version" not in st.session_state:
            st.session_state["qm_version"] = 0
        v = st.session_state["qm_version"]

        # ── Quick Add: same day, same location, multiple technicians ────────
        st.markdown('<div class="sub-hdr">⚡ Quick Add — Same Day &amp; Location, Multiple Technicians</div>', unsafe_allow_html=True)
        qc1, qc2 = st.columns(2)
        with qc1:
            qm_date = st.date_input("Date", value=None, key="qm_date")
        with qc2:
            qm_loc = st.selectbox("Location", ["-- Select --"] + active_locs, key="qm_loc")

        qm_techs = st.multiselect("Technicians who worked today", active_techs, key=f"qm_techs_{v}")

        qty_map = {}
        if qm_techs:
            st.caption("Enter quantities for each technician:")
            for t in qm_techs:
                cc1, cc2, cc3 = st.columns([2, 1, 1])
                with cc1:
                    st.markdown(f"**{t}**")
                with cc2:
                    q1 = st.number_input("1PH", min_value=0, step=1, value=0, key=f"qm_q1_{v}_{t}", label_visibility="collapsed")
                with cc3:
                    q3 = st.number_input("3PH", min_value=0, step=1, value=0, key=f"qm_q3_{v}_{t}", label_visibility="collapsed")
                qty_map[t] = (q1, q3)

        if st.button("➕ Add These To Batch", type="primary", use_container_width=True, disabled=not qm_techs):
            if qm_date is None:
                st.error("❌ Pick a date first.")
            elif qm_loc == "-- Select --":
                st.error("❌ Pick a location first.")
            else:
                added = 0
                for t, (q1, q3) in qty_map.items():
                    if q1 > 0 or q3 > 0:
                        st.session_state["installs_batch"].append({
                            "date": str(qm_date), "tech_name": t, "location": qm_loc,
                            "qty_1ph": int(q1), "qty_3ph": int(q3),
                        })
                        added += 1
                if added:
                    st.session_state["qm_version"] += 1
                    st.success(f"✅ Added {added} entr{'y' if added == 1 else 'ies'} to the batch below.")
                    st.rerun()
                else:
                    st.warning("⚠️ Enter at least one quantity for a selected technician.")

        # ── Single one-off entry (different date/location than the above) ───
        with st.expander("➕ Add a single one-off entry (different date or location)"):
            sc1, sc2 = st.columns(2)
            with sc1:
                single_date = st.date_input("Date", value=None, key=f"single_date_{v}")
            with sc2:
                single_tech = st.selectbox("Technician", ["-- Select --"] + active_techs, key=f"single_tech_{v}")
            single_loc = st.selectbox("Location", ["-- Select --"] + active_locs, key=f"single_loc_{v}")
            sc3, sc4 = st.columns(2)
            with sc3:
                single_q1 = st.number_input("1 PH Qty", min_value=0, step=1, value=0, key=f"single_q1_{v}")
            with sc4:
                single_q3 = st.number_input("3 PH Qty", min_value=0, step=1, value=0, key=f"single_q3_{v}")
            if st.button("➕ Add This Entry To Batch", use_container_width=True):
                if single_date is None or single_tech == "-- Select --" or single_loc == "-- Select --":
                    st.error("❌ Fill date, technician and location.")
                elif single_q1 == 0 and single_q3 == 0:
                    st.error("❌ Enter at least one quantity.")
                else:
                    st.session_state["installs_batch"].append({
                        "date": str(single_date), "tech_name": single_tech, "location": single_loc,
                        "qty_1ph": int(single_q1), "qty_3ph": int(single_q3),
                    })
                    st.session_state["qm_version"] += 1
                    st.success("✅ Added to batch below.")
                    st.rerun()

        # ── Batch preview cards + Save All ───────────────────────────────────
        st.markdown('<div class="sub-hdr">🧾 Batch Ready To Save</div>', unsafe_allow_html=True)
        batch = st.session_state["installs_batch"]
        if not batch:
            st.info("No entries yet — add some above.")
        else:
            for i, entry in enumerate(batch):
                card_col, del_col = st.columns([5, 1])
                with card_col:
                    st.markdown(f"""
                    <div style="background:#ffffff;border:1px solid #E7E9EE;border-radius:12px;
                        padding:10px 14px;margin-bottom:6px;">
                        <b>{entry['tech_name']}</b> — {entry['location']}<br/>
                        <span style="color:#64748b;font-size:.85rem;">
                            {entry['date']} · 1PH: {entry['qty_1ph']} · 3PH: {entry['qty_3ph']}
                        </span>
                    </div>
                    """, unsafe_allow_html=True)
                with del_col:
                    if st.button("🗑️", key=f"del_installs_batch_{i}"):
                        st.session_state["installs_batch"].pop(i)
                        st.rerun()

            bcol1, bcol2 = st.columns(2)
            with bcol1:
                clear_batch = st.button("🗑️ Clear Batch", use_container_width=True)
            with bcol2:
                save_all = st.button(f"💾 Save All ({len(batch)}) To Sheet", type="primary", use_container_width=True)

            if clear_batch:
                st.session_state["installs_batch"] = []
                st.rerun()

            if save_all:
                df_existing = get_data("Installations")
                new_rows, skipped = [], []
                for entry in batch:
                    dup = False
                    if not df_existing.empty and has_col(df_existing, "date", "tech_name"):
                        dup = not df_existing[(df_existing["date"] == entry["date"]) & (df_existing["tech_name"] == entry["tech_name"])].empty
                    if dup:
                        skipped.append(f"{entry['tech_name']} ({entry['date']})")
                    else:
                        new_rows.append({
                            "date": entry["date"], "tech_name": entry["tech_name"], "location": entry["location"],
                            "qty_1ph": str(entry["qty_1ph"]), "qty_3ph": str(entry["qty_3ph"]),
                        })

                if new_rows:
                    updated = pd.concat([df_existing, pd.DataFrame(new_rows)], ignore_index=True) if not df_existing.empty else pd.DataFrame(new_rows)
                    if safe_update("Installations", updated):
                        st.success(f"✅ Saved {len(new_rows)} entr{'y' if len(new_rows) == 1 else 'ies'}.")
                        if skipped:
                            st.warning(f"⚠️ Skipped (already exists for that tech/date): {', '.join(skipped)}")
                        st.session_state["installs_batch"] = []
                        st.rerun()
                else:
                    st.error(f"❌ All entries were duplicates (already exist for that tech/date): {', '.join(skipped)}")

    st.markdown('<div class="sec-hdr">📋 Installation Log</div>', unsafe_allow_html=True)
    log_data = get_data("Installations")

    if log_data.empty:
        st.info("No installation entries yet.")
    else:
        log_sorted = log_data.iloc[::-1].reset_index(drop=True)
        ITEMS = 10
        total_pages = max(1, math.ceil(len(log_sorted) / ITEMS))
        page = st.number_input(f"Page (1 – {total_pages})", min_value=1, max_value=total_pages, step=1, value=1)
        s, e = (page - 1) * ITEMS, page * ITEMS
        disp_log = log_sorted.iloc[s:e].copy()
        if has_col(disp_log, "qty_1ph", "qty_3ph"):
            disp_log["qty_1ph"] = disp_log["qty_1ph"].apply(lambda x: safe_int(x))
            disp_log["qty_3ph"] = disp_log["qty_3ph"].apply(lambda x: safe_int(x))
            disp_log["Total"] = disp_log["qty_1ph"] + disp_log["qty_3ph"]
        st.dataframe(disp_log, use_container_width=True, hide_index=True)

        st.caption("Select a record to edit or delete:")
        log_options_map = {}
        for idx, row in log_sorted.iterrows():
            label = f"#{idx+1}  {row['date']} | {row['tech_name']}"
            log_options_map[label] = idx

        target_label = st.selectbox("Select Record", ["-- Select --"] + list(log_options_map.keys()), key="inst_sel")

        if target_label != "-- Select --":
            sel_idx = log_options_map[target_label]
            curr_row = log_sorted.iloc[sel_idx]
            curr_q1 = safe_int(curr_row.get("qty_1ph", 0))
            curr_q3 = safe_int(curr_row.get("qty_3ph", 0))
            curr_loc = curr_row.get("location", "")
            loc_idx = active_locs.index(curr_loc) if curr_loc in active_locs and active_locs else 0

            st.markdown(f'<div class="warn-box">⚠️ Modifying: <b>{curr_row["tech_name"]}</b> on <b>{curr_row["date"]}</b></div>', unsafe_allow_html=True)
            with st.form("edit_log_form"):
                e_loc = st.selectbox("Location", active_locs, index=loc_idx) if active_locs else st.text_input("Location", value=curr_loc)
                ec1, ec2 = st.columns(2)
                with ec1:
                    e_q1 = st.number_input("1 PH Qty", min_value=0, step=1, value=curr_q1)
                with ec2:
                    e_q3 = st.number_input("3 PH Qty", min_value=0, step=1, value=curr_q3)
                btn_update, btn_delete = st.columns(2)
                with btn_update:
                    do_update = st.form_submit_button("✏️ Update", type="primary")
                with btn_delete:
                    do_delete = st.form_submit_button("🗑️ Delete")

            if do_update:
                if e_q1 == 0 and e_q3 == 0:
                    st.error("❌ Both quantities cannot be 0.")
                else:
                    mask = ((log_data["date"] == curr_row["date"]) & (log_data["tech_name"] == curr_row["tech_name"]))
                    log_data.loc[mask, ["location", "qty_1ph", "qty_3ph"]] = [str(e_loc), str(e_q1), str(e_q3)]
                    if safe_update("Installations", log_data):
                        st.success("✅ Entry updated.")
                        st.rerun()

            if do_delete:
                st.session_state["pending_inst_del"] = curr_row["date"] + "||" + curr_row["tech_name"]

        if "pending_inst_del" in st.session_state:
            del_date, del_tech = st.session_state["pending_inst_del"].split("||", 1)
            st.markdown(f'<div class="warn-box">⚠️ Confirm delete for <b>{del_tech}</b> on <b>{del_date}</b>?</div>', unsafe_allow_html=True)
            cy, cn = st.columns(2)
            with cy:
                if st.button("✅ Yes, Delete", key="conf_del_inst"):
                    mask = ((log_data["date"] == del_date) & (log_data["tech_name"] == del_tech))
                    log_data = log_data[~mask]
                    if safe_update("Installations", log_data):
                        del st.session_state["pending_inst_del"]
                        st.success("Deleted.")
                        st.rerun()
            with cn:
                if st.button("❌ Cancel", key="cancel_del_inst"):
                    del st.session_state["pending_inst_del"]
                    st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
#  INVENTORY (STORE)
# ═══════════════════════════════════════════════════════════════════════════════
with tab_inv:
    st.markdown('<div class="sec-hdr">📥 Inward Store Material</div>', unsafe_allow_html=True)
    with st.form("inv_form", clear_on_submit=True):
        iv1, iv2 = st.columns(2)
        with iv1:
            idate = st.date_input("Received Date", date.today())
            itype = st.selectbox("Type", ["1 PH", "3 PH"])
        with iv2:
            iqty = st.number_input("Quantity", min_value=1, step=1, value=1)
            imrn = st.text_input("MRN No.")
        imake = st.selectbox("Make", ["Schneider", "Genus", "Other"])
        iv_sub = st.form_submit_button("📥 Save Stock", type="primary")

    if iv_sub:
        if not imrn.strip():
            st.error("❌ MRN No. is required.")
        else:
            df_inv_exist = get_data("Inventory")
            new_inv = pd.DataFrame([{"date": str(idate), "type": str(itype), "qty": str(iqty), "mrn": imrn.strip(), "make": str(imake)}])
            updated_inv = pd.concat([df_inv_exist, new_inv], ignore_index=True) if not df_inv_exist.empty else new_inv
            if safe_update("Inventory", updated_inv):
                st.success(f"✅ Inwarded {iqty} × {itype} ({imake}) — MRN {imrn.strip()}")
                st.rerun()

    st.markdown('<div class="sec-hdr">📊 Live Stock Summary</div>', unsafe_allow_html=True)
    df_inv_t = get_data("Inventory")
    df_inst_s = df_installations_master

    r_1ph = r_3ph = u_1ph = u_3ph = 0
    if not df_inv_t.empty and has_col(df_inv_t, "type", "qty"):
        r_1ph = int(safe_numeric_col(df_inv_t[df_inv_t["type"] == "1 PH"], "qty").sum())
        r_3ph = int(safe_numeric_col(df_inv_t[df_inv_t["type"] == "3 PH"], "qty").sum())
    if not df_inst_s.empty and has_col(df_inst_s, "qty_1ph", "qty_3ph"):
        u_1ph = int(safe_numeric_col(df_inst_s, "qty_1ph").sum())
        u_3ph = int(safe_numeric_col(df_inst_s, "qty_3ph").sum())

    sm1, sm2, sm3, sm4 = st.columns(4)
    sm1.metric("1PH Received", r_1ph)
    sm2.metric("3PH Received", r_3ph)
    sm3.metric("1PH Pending Stock", max(r_1ph - u_1ph, 0))
    sm4.metric("3PH Pending Stock", max(r_3ph - u_3ph, 0))

    st.markdown('<div class="sec-hdr">📋 Inventory Log</div>', unsafe_allow_html=True)
    if df_inv_t.empty:
        st.info("No inventory entries yet.")
    else:
        inv_sorted = df_inv_t.iloc[::-1].reset_index(drop=True)
        inv_exp = inv_sorted.rename(columns={"date": "Date", "type": "Type", "qty": "Qty", "mrn": "MRN No", "make": "Make"})
        st.download_button("⬇ Export Inventory CSV", inv_exp.to_csv(index=False).encode(), "inventory.csv", "text/csv", use_container_width=True)

        ITEMS_INV = 10
        total_inv_p = max(1, math.ceil(len(inv_sorted) / ITEMS_INV))
        inv_page = st.number_input(f"Page (1–{total_inv_p})", min_value=1, max_value=total_inv_p, step=1, value=1, key="inv_page")
        si, ei = (inv_page - 1) * ITEMS_INV, inv_page * ITEMS_INV
        st.dataframe(inv_sorted.iloc[si:ei], use_container_width=True, hide_index=True)

        st.caption("Select an inventory entry to edit or delete:")
        inv_options_map = {}
        for idx, row in inv_sorted.iterrows():
            label = f"#{idx+1}  {row.get('date','')} | {row.get('type','')} | MRN:{row.get('mrn','')}"
            inv_options_map[label] = idx

        inv_target = st.selectbox("Select Inventory Record", ["-- Select --"] + list(inv_options_map.keys()), key="inv_sel")

        if inv_target != "-- Select --":
            inv_idx = inv_options_map[inv_target]
            inv_row = inv_sorted.iloc[inv_idx]

            st.markdown(f'<div class="warn-box">⚠️ Modifying: MRN <b>{inv_row.get("mrn","")}</b> — {inv_row.get("type","")} ({inv_row.get("make","")})</div>', unsafe_allow_html=True)
            with st.form("edit_inv_form"):
                ei1, ei2, ei3 = st.columns(3)
                with ei1:
                    e_qty = st.number_input("Quantity", min_value=1, step=1, value=safe_int(inv_row.get("qty", 1), 1))
                with ei2:
                    e_mrn = st.text_input("MRN No.", value=str(inv_row.get("mrn", "")))
                with ei3:
                    make_opts = ["Schneider", "Genus", "Other"]
                    curr_make = inv_row.get("make", "Schneider")
                    mk_idx = make_opts.index(curr_make) if curr_make in make_opts else 0
                    e_make = st.selectbox("Make", make_opts, index=mk_idx)
                ib1, ib2 = st.columns(2)
                with ib1:
                    inv_do_update = st.form_submit_button("✏️ Update", type="primary")
                with ib2:
                    inv_do_delete = st.form_submit_button("🗑️ Delete")

            if inv_do_update:
                if not e_mrn.strip():
                    st.error("❌ MRN No. cannot be empty.")
                else:
                    orig_df = df_inv_t.copy()
                    orig_inv_idx = len(orig_df) - 1 - inv_idx
                    orig_df.iloc[orig_inv_idx, orig_df.columns.get_loc("qty")] = str(e_qty)
                    orig_df.iloc[orig_inv_idx, orig_df.columns.get_loc("mrn")] = e_mrn.strip()
                    orig_df.iloc[orig_inv_idx, orig_df.columns.get_loc("make")] = e_make
                    if safe_update("Inventory", orig_df):
                        st.success("✅ Inventory entry updated.")
                        st.rerun()

            if inv_do_delete:
                st.session_state["pending_inv_del"] = inv_idx

        if "pending_inv_del" in st.session_state:
            del_inv_idx = st.session_state["pending_inv_del"]
            st.markdown('<div class="warn-box">⚠️ Confirm delete? This will affect stock totals.</div>', unsafe_allow_html=True)
            iy, inv_n = st.columns(2)
            with iy:
                if st.button("✅ Yes, Delete", key="conf_del_inv"):
                    orig_df = df_inv_t.copy()
                    orig_inv_ri = len(orig_df) - 1 - del_inv_idx
                    orig_df = orig_df.drop(index=orig_inv_ri).reset_index(drop=True)
                    if safe_update("Inventory", orig_df):
                        del st.session_state["pending_inv_del"]
                        st.success("Deleted.")
                        st.rerun()
            with inv_n:
                if st.button("❌ Cancel", key="cancel_del_inv"):
                    del st.session_state["pending_inv_del"]
                    st.rerun()

# ═══════════════════════════════════════════════════════════════════════════════
#  ADMIN
# ═══════════════════════════════════════════════════════════════════════════════
with tab_admin:
    st.markdown("""
    <div class="warn-box" style="background:#f8f9fa;border-color:#cbd5e1;color:#475569;">
    💡 <b>Tip:</b> Add one or several at once below, review them as cards, then Save Batch.
    Existing entries are listed further down as cards — tap ✏️ Edit to change details or toggle
    Active/Inactive, or 🗑️ to delete.
    </div>
    """, unsafe_allow_html=True)

    subtab_tech, subtab_loc = st.tabs(["👷 Technicians", "📍 Locations"])

    # ── Technicians ───────────────────────────────────────────────────────────
    with subtab_tech:
        if "tech_batch" not in st.session_state:
            st.session_state["tech_batch"] = []
        if "tech_form_version" not in st.session_state:
            st.session_state["tech_form_version"] = 0
        tv = st.session_state["tech_form_version"]

        st.markdown('<div class="sub-hdr">➕ Add Technicians (one or several)</div>', unsafe_allow_html=True)
        st.caption("Login ID is optional — set it to match the 'Installer LoginID' column (e.g. TL_Vinod) in the MDM export so bulk uploads auto-map to this technician's name.")
        tc1, tc2, tc3, tc4 = st.columns([2, 1, 1, 1.3])
        with tc1:
            new_t_name = st.text_input("Name", key=f"new_t_name_{tv}")
        with tc2:
            new_t_phone = st.text_input("Phone (optional)", key=f"new_t_phone_{tv}")
        with tc3:
            new_t_aadhar = st.text_input("Aadhar (optional)", key=f"new_t_aadhar_{tv}")
        with tc4:
            new_t_login = st.text_input("Login ID (optional)", key=f"new_t_login_{tv}", placeholder="TL_Vinod")

        if st.button("➕ Add To Batch", key="add_tech_batch_btn", type="primary", use_container_width=True):
            if not new_t_name.strip():
                st.error("❌ Name is required.")
            elif any(b["name"] == new_t_name.strip() for b in st.session_state["tech_batch"]):
                st.error("❌ Already added to this batch.")
            else:
                st.session_state["tech_batch"].append({
                    "name": new_t_name.strip(), "phone": new_t_phone.strip(),
                    "aadhar": new_t_aadhar.strip(), "login_id": new_t_login.strip(),
                })
                st.session_state["tech_form_version"] += 1
                st.rerun()

        if st.session_state["tech_batch"]:
            st.markdown('<div class="sub-hdr">🧾 Batch Ready To Save</div>', unsafe_allow_html=True)
            for i, b in enumerate(st.session_state["tech_batch"]):
                bcard, bdel = st.columns([5, 1])
                with bcard:
                    detail = " · ".join([x for x in [b["phone"], b["aadhar"], b.get("login_id", "")] if x]) or "no phone/aadhar/login given"
                    st.markdown(f"""
                    <div style="background:#ffffff;border:1px solid #E7E9EE;border-radius:12px;padding:10px 14px;margin-bottom:6px;">
                        <b>{b['name']}</b><br/><span style="color:#64748b;font-size:.85rem;">{detail}</span>
                    </div>
                    """, unsafe_allow_html=True)
                with bdel:
                    if st.button("🗑️", key=f"del_tech_batch_{i}"):
                        st.session_state["tech_batch"].pop(i)
                        st.rerun()

            if st.button(f"💾 Save Batch ({len(st.session_state['tech_batch'])})", key="save_tech_batch", type="primary", use_container_width=True):
                df_t_exist = get_data("Technicians")
                existing_names = set(df_t_exist["name"].values) if (not df_t_exist.empty and "name" in df_t_exist.columns) else set()
                new_rows, skipped = [], []
                for b in st.session_state["tech_batch"]:
                    if b["name"] in existing_names:
                        skipped.append(b["name"])
                    else:
                        new_rows.append({"name": b["name"], "phone": b["phone"], "aadhar": b["aadhar"], "is_active": "1", "login_id": b.get("login_id", "")})
                if new_rows:
                    updated = pd.concat([df_t_exist, pd.DataFrame(new_rows)], ignore_index=True) if not df_t_exist.empty else pd.DataFrame(new_rows)
                    if safe_update("Technicians", updated):
                        st.success(f"✅ Added {len(new_rows)} technician(s).")
                        if skipped:
                            st.warning(f"⚠️ Skipped (already exist): {', '.join(skipped)}")
                        st.session_state["tech_batch"] = []
                        st.rerun()
                else:
                    st.error(f"❌ All names already exist: {', '.join(skipped)}")

        st.markdown('<div class="sec-hdr">👷 Existing Technicians</div>', unsafe_allow_html=True)
        df_t = df_technicians_master.copy()
        if not df_t.empty:
            df_t = df_t.rename(columns={c: str(c).strip().lower() for c in df_t.columns})
            for col in ["name", "phone", "aadhar", "is_active", "login_id"]:
                if col not in df_t.columns:
                    df_t[col] = ""

        if df_t.empty:
            st.info("No technicians added yet.")
        else:
            for idx, row in df_t.iterrows():
                is_active = str(row.get("is_active", "1")).strip() in ["1", "1.0", "true", "yes"]
                pill_color = "#0B7A56" if is_active else "#94a3b8"
                pill_bg = "#E6F7F0" if is_active else "#f1f5f9"
                pill_text = "Active" if is_active else "Inactive"

                rc1, rc2 = st.columns([5, 2])
                with rc1:
                    detail = " · ".join([x for x in [str(row.get("phone", "")), str(row.get("aadhar", "")), (f"Login: {row.get('login_id','')}" if str(row.get("login_id","")).strip() else "")] if x]) or "no phone/aadhar/login on file"
                    st.markdown(f"""
                    <div style="background:#ffffff;border:1px solid #E7E9EE;border-radius:12px;padding:10px 14px;margin-bottom:6px;">
                        <b>{row.get('name','')}</b>
                        <span style="background:{pill_bg};color:{pill_color};border-radius:20px;padding:2px 10px;
                            font-size:.72rem;font-weight:700;margin-left:8px;">{pill_text}</span><br/>
                        <span style="color:#64748b;font-size:.85rem;">{detail}</span>
                    </div>
                    """, unsafe_allow_html=True)
                with rc2:
                    ecol, dcol = st.columns(2)
                    with ecol:
                        edit_clicked = st.button("✏️", key=f"edit_tech_{idx}")
                    with dcol:
                        del_clicked = st.button("🗑️", key=f"del_tech_{idx}")

                if edit_clicked:
                    st.session_state["editing_tech_idx"] = idx
                if del_clicked:
                    st.session_state["deleting_tech_idx"] = idx

                if st.session_state.get("editing_tech_idx") == idx:
                    with st.form(f"edit_tech_form_{idx}"):
                        e_name = st.text_input("Name", value=str(row.get("name", "")))
                        e_phone = st.text_input("Phone (optional)", value=str(row.get("phone", "")))
                        e_aadhar = st.text_input("Aadhar (optional)", value=str(row.get("aadhar", "")))
                        e_login = st.text_input("Login ID (optional)", value=str(row.get("login_id", "")), placeholder="TL_Vinod")
                        e_active = st.selectbox("Status", ["Active", "Inactive"], index=0 if is_active else 1)
                        sv, cn = st.columns(2)
                        with sv:
                            do_save = st.form_submit_button("💾 Save", type="primary")
                        with cn:
                            do_cancel = st.form_submit_button("Cancel")
                    if do_save:
                        if not e_name.strip():
                            st.error("❌ Name cannot be empty.")
                        else:
                            df_t.loc[idx, ["name", "phone", "aadhar", "login_id", "is_active"]] = [
                                e_name.strip(), e_phone.strip(), e_aadhar.strip(), e_login.strip(), "1" if e_active == "Active" else "0"
                            ]
                            if safe_update("Technicians", df_t):
                                del st.session_state["editing_tech_idx"]
                                st.success("✅ Updated.")
                                st.rerun()
                    if do_cancel:
                        del st.session_state["editing_tech_idx"]
                        st.rerun()

                if st.session_state.get("deleting_tech_idx") == idx:
                    st.markdown(f'<div class="warn-box">⚠️ Delete <b>{row.get("name","")}</b>? This removes them from future entry forms.</div>', unsafe_allow_html=True)
                    yc, ncol = st.columns(2)
                    with yc:
                        if st.button("✅ Yes, Delete", key=f"conf_del_tech_{idx}"):
                            df_t_new = df_t.drop(index=idx).reset_index(drop=True)
                            if safe_update("Technicians", df_t_new):
                                del st.session_state["deleting_tech_idx"]
                                st.success("Deleted.")
                                st.rerun()
                    with ncol:
                        if st.button("❌ Cancel", key=f"cancel_del_tech_{idx}"):
                            del st.session_state["deleting_tech_idx"]
                            st.rerun()

    # ── Locations ─────────────────────────────────────────────────────────────
    with subtab_loc:
        if "loc_batch" not in st.session_state:
            st.session_state["loc_batch"] = []
        if "loc_form_version" not in st.session_state:
            st.session_state["loc_form_version"] = 0
        lv = st.session_state["loc_form_version"]

        st.markdown('<div class="sub-hdr">➕ Add Locations (one or several)</div>', unsafe_allow_html=True)
        new_loc_name = st.text_input("Location Name", key=f"new_loc_name_{lv}")

        if st.button("➕ Add To Batch", key="add_loc_batch_btn", type="primary", use_container_width=True):
            if not new_loc_name.strip():
                st.error("❌ Location name is required.")
            elif new_loc_name.strip() in st.session_state["loc_batch"]:
                st.error("❌ Already added to this batch.")
            else:
                st.session_state["loc_batch"].append(new_loc_name.strip())
                st.session_state["loc_form_version"] += 1
                st.rerun()

        if st.session_state["loc_batch"]:
            st.markdown('<div class="sub-hdr">🧾 Batch Ready To Save</div>', unsafe_allow_html=True)
            for i, l in enumerate(st.session_state["loc_batch"]):
                bcard, bdel = st.columns([5, 1])
                with bcard:
                    st.markdown(f"""
                    <div style="background:#ffffff;border:1px solid #E7E9EE;border-radius:12px;padding:10px 14px;margin-bottom:6px;">
                        <b>{l}</b>
                    </div>
                    """, unsafe_allow_html=True)
                with bdel:
                    if st.button("🗑️", key=f"del_loc_batch_{i}"):
                        st.session_state["loc_batch"].pop(i)
                        st.rerun()

            if st.button(f"💾 Save Batch ({len(st.session_state['loc_batch'])})", key="save_loc_batch", type="primary", use_container_width=True):
                df_l_exist = get_data("Locations")
                existing_locs = set(df_l_exist["location_name"].values) if (not df_l_exist.empty and "location_name" in df_l_exist.columns) else set()
                new_rows, skipped = [], []
                for l in st.session_state["loc_batch"]:
                    if l in existing_locs:
                        skipped.append(l)
                    else:
                        new_rows.append({"location_name": l})
                if new_rows:
                    updated = pd.concat([df_l_exist, pd.DataFrame(new_rows)], ignore_index=True) if not df_l_exist.empty else pd.DataFrame(new_rows)
                    if safe_update("Locations", updated):
                        st.success(f"✅ Added {len(new_rows)} location(s).")
                        if skipped:
                            st.warning(f"⚠️ Skipped (already exist): {', '.join(skipped)}")
                        st.session_state["loc_batch"] = []
                        st.rerun()
                else:
                    st.error(f"❌ All locations already exist: {', '.join(skipped)}")

        st.markdown('<div class="sec-hdr">📍 Existing Locations</div>', unsafe_allow_html=True)
        df_l = df_locations_master.copy()
        if not df_l.empty:
            df_l = df_l.rename(columns={c: str(c).strip().lower() for c in df_l.columns})
            if "location_name" not in df_l.columns:
                df_l["location_name"] = ""

        if df_l.empty:
            st.info("No locations added yet.")
        else:
            for idx, row in df_l.iterrows():
                rc1, rc2 = st.columns([5, 2])
                with rc1:
                    st.markdown(f"""
                    <div style="background:#ffffff;border:1px solid #E7E9EE;border-radius:12px;padding:10px 14px;margin-bottom:6px;">
                        <b>{row.get('location_name','')}</b>
                    </div>
                    """, unsafe_allow_html=True)
                with rc2:
                    ecol, dcol = st.columns(2)
                    with ecol:
                        edit_loc_clicked = st.button("✏️", key=f"edit_loc_{idx}")
                    with dcol:
                        del_loc_clicked = st.button("🗑️", key=f"del_loc_{idx}")

                if edit_loc_clicked:
                    st.session_state["editing_loc_idx"] = idx
                if del_loc_clicked:
                    st.session_state["deleting_loc_idx"] = idx

                if st.session_state.get("editing_loc_idx") == idx:
                    with st.form(f"edit_loc_form_{idx}"):
                        e_loc_name = st.text_input("Location Name", value=str(row.get("location_name", "")))
                        sv, cn = st.columns(2)
                        with sv:
                            do_save_loc = st.form_submit_button("💾 Save", type="primary")
                        with cn:
                            do_cancel_loc = st.form_submit_button("Cancel")
                    if do_save_loc:
                        if not e_loc_name.strip():
                            st.error("❌ Location name cannot be empty.")
                        else:
                            df_l.loc[idx, "location_name"] = e_loc_name.strip()
                            if safe_update("Locations", df_l):
                                del st.session_state["editing_loc_idx"]
                                st.success("✅ Updated.")
                                st.rerun()
                    if do_cancel_loc:
                        del st.session_state["editing_loc_idx"]
                        st.rerun()

                if st.session_state.get("deleting_loc_idx") == idx:
                    st.markdown(f'<div class="warn-box">⚠️ Delete <b>{row.get("location_name","")}</b>?</div>', unsafe_allow_html=True)
                    yc, ncol = st.columns(2)
                    with yc:
                        if st.button("✅ Yes, Delete", key=f"conf_del_loc_{idx}"):
                            df_l_new = df_l.drop(index=idx).reset_index(drop=True)
                            if safe_update("Locations", df_l_new):
                                del st.session_state["deleting_loc_idx"]
                                st.success("Deleted.")
                                st.rerun()
                    with ncol:
                        if st.button("❌ Cancel", key=f"cancel_del_loc_{idx}"):
                            del st.session_state["deleting_loc_idx"]
                            st.rerun()
